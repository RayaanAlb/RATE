#!/usr/bin/env python3
"""
Command-line QR Code Generator
Generates QR codes with Serial Number, Verification Code, and DevUID
Saves data to SQLite database
"""

import qrcode
import sqlite3
from datetime import datetime
import sys
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Alignment, PatternFill
from PIL import Image

class QRGeneratorCLI:
    def __init__(self):
        self.init_database()
    
    def init_database(self):
        """Initialize SQLite database"""
        self.conn = sqlite3.connect('qr_codes.db')
        self.cursor = self.conn.cursor()
        
        # Check if we need to migrate from auto-increment to manual ID management
        self.cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='qr_records'")
        table_exists = self.cursor.fetchone()
        
        if table_exists:
            # Check if table has auto-increment
            self.cursor.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='qr_records'")
            table_sql = self.cursor.fetchone()[0]
            
            if 'AUTOINCREMENT' in table_sql:
                print("🔄 Migrating database to remove auto-increment...")
                self.migrate_database()
        else:
            # Create new table without auto-increment
            self.cursor.execute('''
                CREATE TABLE qr_records (
                    id INTEGER PRIMARY KEY,
                    serial_number TEXT NOT NULL,
                    verification_code TEXT NOT NULL,
                    dev_uid TEXT NOT NULL,
                    qr_filename TEXT NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            self.conn.commit()
    
    def migrate_database(self):
        """Migrate existing database to remove auto-increment"""
        try:
            # Create backup table
            self.cursor.execute('''
                CREATE TABLE qr_records_backup AS 
                SELECT * FROM qr_records ORDER BY created_at ASC
            ''')
            
            # Drop original table
            self.cursor.execute('DROP TABLE qr_records')
            
            # Create new table without auto-increment
            self.cursor.execute('''
                CREATE TABLE qr_records (
                    id INTEGER PRIMARY KEY,
                    serial_number TEXT NOT NULL,
                    verification_code TEXT NOT NULL,
                    dev_uid TEXT NOT NULL,
                    qr_filename TEXT NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # Get data from backup and reassign sequential IDs
            self.cursor.execute('SELECT serial_number, verification_code, dev_uid, qr_filename, created_at FROM qr_records_backup ORDER BY created_at ASC')
            records = self.cursor.fetchall()
            
            # Insert records with sequential IDs starting from 1
            for i, record in enumerate(records, 1):
                self.cursor.execute('''
                    INSERT INTO qr_records (id, serial_number, verification_code, dev_uid, qr_filename, created_at)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (i, record[0], record[1], record[2], record[3], record[4]))
            
            # Drop backup table
            self.cursor.execute('DROP TABLE qr_records_backup')
            
            self.conn.commit()
            print("✅ Database migration completed - IDs are now sequential")
            
        except Exception as e:
            print(f"❌ Migration error: {str(e)}")
            # Restore from backup if migration fails
            try:
                self.cursor.execute('DROP TABLE IF EXISTS qr_records')
                self.cursor.execute('ALTER TABLE qr_records_backup RENAME TO qr_records')
                self.conn.commit()
                print("🔄 Restored original table")
            except:
                pass
    
    def generate_qr_code(self, serial_number, verification_code, dev_uid, format_type="olarm"):
        """Generate QR code with input data in specified format"""
        try:
            # Create QR code
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_M,  # Changed to Medium for better reliability
                box_size=10,
                border=4,
            )
            
            # Prepare QR code data based on format type
            if format_type == "json":
                import json
                qr_data = json.dumps({
                    "sn": serial_number,
                    "vc": verification_code,
                    "uid": dev_uid
                }, separators=(',', ':'))  # Compact JSON
            elif format_type == "csv":
                qr_data = f"{serial_number},{verification_code},{dev_uid}"
            elif format_type == "pipe":
                qr_data = f"{serial_number}|{verification_code}|{dev_uid}"
            elif format_type == "labeled":
                # Old format with labels
                qr_data = f"Serial Number: {serial_number}\nVerification Code: {verification_code}\nDevUID: {dev_uid}"
            elif format_type == "compact":
                # Very compact format
                qr_data = f"{serial_number}:{verification_code}:{dev_uid}"
            elif format_type == "url":
                # URL format for web validation (generic)
                qr_data = f"https://validate.example.com?sn={serial_number}&vc={verification_code}&uid={dev_uid}"
            elif format_type == "olarm":
                # Olarm specific URL format: serial,devuid,verification_code
                qr_data = f"https://olarm.com/o/flxr?a={serial_number},{dev_uid},{verification_code}"
            else:
                # Default to JSON if invalid format specified
                import json
                qr_data = json.dumps({
                    "sn": serial_number,
                    "vc": verification_code,
                    "uid": dev_uid
                }, separators=(',', ':'))
            
            qr.add_data(qr_data)
            qr.make(fit=True)
            
            # Create QR code image
            qr_image = qr.make_image(fill_color="black", back_color="white")
            
            # Generate filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"qr_code_{serial_number}_{timestamp}.png"
            
            # Save QR code image
            qr_image.save(filename)
            
            # Save to database
            self.save_to_database(serial_number, verification_code, dev_uid, filename)
            
            # Auto-export to Excel after each new record
            self.export_to_excel(verbose=False)
            
            print(f"✅ QR code generated successfully!")
            print(f"📁 Saved as: {filename}")
            print(f"🔧 Format: {format_type}")
            print(f"📋 Content: {qr_data}")
            print(f"💾 Record saved to database")
            
            return filename
            
        except Exception as e:
            print(f"❌ Error generating QR code: {str(e)}")
            return None
    
    def save_to_database(self, serial_number, verification_code, dev_uid, filename):
        """Save record to database"""
        try:
            # Find the lowest available ID (reuse deleted numbers)
            # First check if there are any records
            self.cursor.execute('SELECT COUNT(*) FROM qr_records')
            record_count = self.cursor.fetchone()[0]
            
            if record_count == 0:
                next_id = 1
            else:
                # Look for gaps in the sequence starting from 1
                self.cursor.execute('SELECT id FROM qr_records ORDER BY id')
                existing_ids = [row[0] for row in self.cursor.fetchall()]
                
                # Find the first missing number in sequence
                next_id = 1
                for existing_id in existing_ids:
                    if next_id == existing_id:
                        next_id += 1
                    else:
                        break
            
            # Insert with specific ID
            self.cursor.execute('''
                INSERT INTO qr_records (id, serial_number, verification_code, dev_uid, qr_filename)
                VALUES (?, ?, ?, ?, ?)
            ''', (next_id, serial_number, verification_code, dev_uid, filename))
            self.conn.commit()
            
            print(f"💾 Record saved with ID: {next_id}")
            
        except Exception as e:
            print(f"❌ Database error: {str(e)}")
    
    def view_records(self, limit=10):
        """View recent records from database"""
        try:
            self.cursor.execute('''
                SELECT id, serial_number, verification_code, dev_uid, qr_filename, created_at
                FROM qr_records
                ORDER BY created_at ASC
                LIMIT ?
            ''', (limit,))
            records = self.cursor.fetchall()
            
            if not records:
                print("📋 No records found.")
                return
            
            print(f"\n📋 Records (Oldest to Newest - Showing {len(records)}):")
            print("-" * 88)
            print(f"{'#':<3} {'ID':<4} {'Serial':<15} {'VCode':<12} {'DevUID':<16} {'Created':<19}")
            print("-" * 88)
            
            for i, record in enumerate(records, 1):
                created_at = datetime.fromisoformat(record[5]).strftime("%Y-%m-%d %H:%M:%S")
                print(f"{i:<3} {record[0]:<4} {record[1]:<15} {record[2]:<12} {record[3]:<16} {created_at}")
            
        except Exception as e:
            print(f"❌ Error loading records: {str(e)}")
    
    def remove_record(self):
        """Remove a record from the database"""
        try:
            # First, show all records
            self.cursor.execute('''
                SELECT id, serial_number, verification_code, dev_uid, qr_filename, created_at
                FROM qr_records
                ORDER BY created_at ASC
            ''')
            records = self.cursor.fetchall()
            
            if not records:
                print("📋 No records found to remove.")
                return
            
            print(f"\n📋 All Records:")
            print("-" * 98)
            print(f"{'#':<3} {'ID':<4} {'Serial':<15} {'VCode':<12} {'DevUID':<16} {'QR File':<25} {'Created':<19}")
            print("-" * 98)
            
            for i, record in enumerate(records, 1):
                created_at = datetime.fromisoformat(record[5]).strftime("%Y-%m-%d %H:%M:%S")
                qr_file = os.path.basename(record[4])  # Show just filename
                print(f"{i:<3} {record[0]:<4} {record[1]:<15} {record[2]:<12} {record[3]:<16} {qr_file:<25} {created_at}")
            
            print("\n🗑️  Enter the ID of the record to remove (or 0 to cancel):")
            record_id = input("Record ID: ").strip()
            
            try:
                record_id = int(record_id)
                if record_id == 0:
                    print("❌ Operation cancelled.")
                    return
                
                # Check if record exists
                self.cursor.execute('SELECT id, qr_filename FROM qr_records WHERE id = ?', (record_id,))
                record = self.cursor.fetchone()
                
                if not record:
                    print(f"❌ Record with ID {record_id} not found.")
                    return
                
                # Confirm deletion
                confirm = input(f"⚠️  Are you sure you want to delete record ID {record_id}? (y/N): ").strip().lower()
                if confirm != 'y':
                    print("❌ Operation cancelled.")
                    return
                
                # Delete the QR code file if it exists
                qr_filename = record[1]
                if os.path.exists(qr_filename):
                    try:
                        os.remove(qr_filename)
                        print(f"🗑️  Deleted QR code file: {qr_filename}")
                    except Exception as e:
                        print(f"⚠️  Warning: Could not delete QR code file {qr_filename}: {e}")
                
                # Delete from database
                self.cursor.execute('DELETE FROM qr_records WHERE id = ?', (record_id,))
                self.conn.commit()
                
                print(f"✅ Record ID {record_id} deleted successfully!")
                
                # Update Excel file
                self.export_to_excel(verbose=False)
                
            except ValueError:
                print("❌ Invalid ID. Please enter a number.")
                
        except Exception as e:
            print(f"❌ Error removing record: {str(e)}")
    
    def export_to_excel(self, verbose=True):
        """Export all records to Excel with QR code images"""
        try:
            # Fetch all records from database
            self.cursor.execute('''
                SELECT id, serial_number, verification_code, dev_uid, qr_filename, created_at
                FROM qr_records
                ORDER BY created_at ASC
            ''')
            records = self.cursor.fetchall()
            
            if not records:
                print("📋 No records found to export.")
                return
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "QR Code Records"
            
            # Define headers
            headers = ['#', 'ID', 'Serial Number', 'Verification Code', 'DevUID', 'QR Code', 'Created At']
            
            # Style for headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Set column widths
            column_widths = [5, 8, 20, 18, 20, 25, 22]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
            
            # Write data and embed QR codes
            for row_idx, record in enumerate(records, 2):
                # Write sequential number and data
                ws.cell(row=row_idx, column=1, value=row_idx - 1)  # Sequential number (#)
                ws.cell(row=row_idx, column=2, value=record[0])  # ID
                ws.cell(row=row_idx, column=3, value=record[1])  # Serial Number
                ws.cell(row=row_idx, column=4, value=record[2])  # Verification Code
                ws.cell(row=row_idx, column=5, value=record[3])  # DevUID
                
                # Format created_at
                created_at = datetime.fromisoformat(record[5]).strftime("%Y-%m-%d %H:%M:%S")
                ws.cell(row=row_idx, column=7, value=created_at)  # Created At
                
                # Add QR code image if file exists
                qr_filename = record[4]
                if os.path.exists(qr_filename):
                    try:
                        # Insert image into Excel directly
                        img_excel = OpenpyxlImage(qr_filename)
                        
                        # Resize the image in Excel (width and height in points)
                        img_excel.width = 120  # approximately 150 pixels
                        img_excel.height = 120  # approximately 150 pixels
                        
                        # Position the image in the QR Code column (now column F)
                        img_excel.anchor = f"F{row_idx}"
                        ws.add_image(img_excel)
                        
                        # Set row height to accommodate image (in points)
                        ws.row_dimensions[row_idx].height = 90
                        
                    except Exception as e:
                        ws.cell(row=row_idx, column=6, value=f"Error loading: {qr_filename}")
                        print(f"⚠️  Warning: Could not load QR code image {qr_filename}: {e}")
                else:
                    ws.cell(row=row_idx, column=6, value=f"File not found: {qr_filename}")
            
            # Use fixed filename that gets overwritten
            excel_filename = "qr_records.xlsx"
            
            # Save workbook
            wb.save(excel_filename)
            
            if verbose:
                print(f"✅ Excel export completed successfully!")
                print(f"📁 Saved as: {excel_filename}")
                print(f"📊 Exported {len(records)} records with QR code images")
                print(f"📝 File will be overwritten on next export")
            else:
                print(f"📊 Excel file updated: {excel_filename}")
            
        except Exception as e:
            print(f"❌ Error exporting to Excel: {str(e)}")
    
    def interactive_mode(self):
        """Interactive mode for generating QR codes"""
        print("🔄 Interactive QR Code Generator")
        print("=" * 40)
        
        while True:
            print("\nOptions:")
            print("1. Generate new QR code")
            print("2. View recent records")
            print("3. Remove a record")
            print("4. Export to Excel")
            print("5. Exit")
            
            choice = input("\nEnter your choice (1-5): ").strip()
            
            if choice == '1':
                print("\n📝 Enter QR Code Information:")
                serial_number = input("Serial Number: ").strip()
                verification_code = input("Verification Code: ").strip()
                dev_uid = input("DevUID: ").strip()
                
                if not all([serial_number, verification_code, dev_uid]):
                    print("❌ Error: All fields are required!")
                    continue
                
                print("\n🔧 Select QR Code Format:")
                print("1. JSON (recommended) - {\"sn\":\"...\",\"vc\":\"...\",\"uid\":\"...\"}")
                print("2. CSV - serial,verification,devuid")
                print("3. Pipe - serial|verification|devuid")
                print("4. Compact - serial:verification:devuid")
                print("5. Labeled (old format) - Serial Number: ... etc")
                print("6. URL - https://validate.example.com?sn=...&vc=...&uid=...")
                print("7. Olarm (recommended) - https://olarm.com/o/flxr?a=serial,devuid,vcode")
                
                format_choice = input("Format choice (1-7, default 7): ").strip()
                format_map = {
                    '1': 'json', '2': 'csv', '3': 'pipe', 
                    '4': 'compact', '5': 'labeled', '6': 'url', '7': 'olarm'
                }
                format_type = format_map.get(format_choice, 'olarm')
                
                self.generate_qr_code(serial_number, verification_code, dev_uid, format_type)
                
            elif choice == '2':
                limit = input("Number of records to show (default 10): ").strip()
                try:
                    limit = int(limit) if limit else 10
                except ValueError:
                    limit = 10
                self.view_records(limit)
                
            elif choice == '3':
                print("\n🗑️  Remove Record")
                self.remove_record()
                
            elif choice == '4':
                print("\n📊 Exporting records to Excel...")
                self.export_to_excel()
                
            elif choice == '5':
                print("👋 Goodbye!")
                break
                
            else:
                print("❌ Invalid choice. Please enter 1, 2, 3, 4, or 5.")
    
    def __del__(self):
        """Close database connection"""
        if hasattr(self, 'conn'):
            self.conn.close()

def main():
    print("🔲 QR Code Generator (Command Line)")
    print("=" * 40)
    
    generator = QRGeneratorCLI()
    
    # Check if command line arguments are provided
    if len(sys.argv) == 4:
        # Command line mode
        serial_number = sys.argv[1]
        verification_code = sys.argv[2]
        dev_uid = sys.argv[3]
        
        print(f"📝 Generating QR code for:")
        print(f"   Serial Number: {serial_number}")
        print(f"   Verification Code: {verification_code}")
        print(f"   DevUID: {dev_uid}")
        print()
        
        generator.generate_qr_code(serial_number, verification_code, dev_uid)
        
    else:
        # Interactive mode
        if len(sys.argv) > 1:
            print("❌ Usage: python3 qr_generator_cli.py [serial_number] [verification_code] [dev_uid]")
            print("   Or run without arguments for interactive mode")
            sys.exit(1)
        
        generator.interactive_mode()

if __name__ == "__main__":
    main() 