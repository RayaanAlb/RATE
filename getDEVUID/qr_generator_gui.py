import flet as ft
import qrcode
import sqlite3
from datetime import datetime, timezone, timedelta
from PIL import Image
import os
import json
import random
import base64
import io
import subprocess
import tempfile
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Alignment, PatternFill

class QRGeneratorApp:
    def __init__(self, page: ft.Page):
        self.page = page
        self.page.title = "QR Code Generator"
        
        # Define South African timezone (SAST = GMT+2)
        self.sast_tz = timezone(timedelta(hours=2))
        
        # Make window resizable and responsive with better minimum sizes
        self.page.window_width = 1200
        self.page.window_height = 850
        self.page.window_min_width = 800
        self.page.window_min_height = 600
        self.page.window_resizable = True
        self.page.window_maximizable = True
        
        self.page.theme_mode = ft.ThemeMode.LIGHT
        self.page.padding = ft.padding.all(10)
        self.page.scroll = ft.ScrollMode.AUTO
        
        # Initialize database in main thread
        self.init_database()
        
        # Create UI controls
        self.create_controls()
        
        # Load existing records
        self.load_records()
        
        # Build the page
        self.build()
    
    def init_database(self):
        """Initialize SQLite database"""
        self.conn = sqlite3.connect('qr_codes.db', check_same_thread=False)
        self.cursor = self.conn.cursor()
        
        # Create table if it doesn't exist
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS qr_records (
                id INTEGER PRIMARY KEY,
                serial_number TEXT NOT NULL,
                verification_code TEXT NOT NULL,
                dev_uid TEXT NOT NULL,
                device_name TEXT,
                qr_filename TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT (datetime('now', '+2 hours'))
            )
        ''')
        
        # Add device_name column if it doesn't exist (for existing databases)
        try:
            self.cursor.execute("ALTER TABLE qr_records ADD COLUMN device_name TEXT")
            self.conn.commit()
            print("✅ Added device_name column to existing database")
        except sqlite3.OperationalError:
            # Column already exists, ignore
            pass
        
        self.conn.commit()
    
    def create_controls(self):
        """Create Flet UI controls"""
        # Title - properly centered for all screen sizes
        self.title = ft.Container(
            content=ft.Text(
                "🔲 QR Code Generator",
                size=28,
                weight=ft.FontWeight.BOLD,
                text_align=ft.TextAlign.CENTER,
                color=ft.Colors.BLUE_700
            ),
            alignment=ft.alignment.center,
            expand=True
        )
        
        # Input fields - responsive width
        self.serial_field = ft.TextField(
            label="Serial Number",
            expand=True,
            autofocus=True
        )
        
        self.vcode_field = ft.TextField(
            label="Verification Code",
            expand=True
        )
        
        self.devuid_field = ft.TextField(
            label="DevUID",
            expand=True
        )
        
        # Get DevUID button
        self.get_devuid_btn = ft.OutlinedButton(
            "🔌 Get DevUID",
            on_click=self.get_devuid_from_device,
            tooltip="Extract DevUID from connected STM32 device via ST-Link",
            style=ft.ButtonStyle(
                text_style=ft.TextStyle(
                    size=12,
                    weight=ft.FontWeight.NORMAL
                ),
                alignment=ft.alignment.center,
                padding=ft.padding.symmetric(horizontal=8, vertical=8),
                shape=ft.RoundedRectangleBorder(radius=8)
            )
        )
        
        self.device_name_field = ft.TextField(
            label="Device Name (Optional)",
            expand=True,
            hint_text="e.g., Living Room Sensor"
        )
        
        # Format dropdown - responsive
        self.format_dropdown = ft.Dropdown(
            label="QR Format",
            expand=True,
            value="olarm",
            options=[
                ft.dropdown.Option("olarm", "Olarm (recommended)"),
                ft.dropdown.Option("json", "JSON"),
                ft.dropdown.Option("csv", "CSV"),
                ft.dropdown.Option("pipe", "Pipe-separated"),
                ft.dropdown.Option("compact", "Compact"),
                ft.dropdown.Option("labeled", "Labeled"),
                ft.dropdown.Option("url", "URL")
            ],
            on_change=self.on_format_change
        )
        
        # Format help text
        self.format_help = ft.Text(
            "Olarm (recommended) - Compatible with Olarm validation system",
            size=12,
            color=ft.Colors.GREY,
            italic=True
        )
        
        # Buttons - responsive with adaptive height and centered text for all window sizes
        self.generate_btn = ft.ElevatedButton(
            "Generate QR Code",
            on_click=self.generate_qr_code,
            bgcolor=ft.Colors.BLUE,
            color=ft.Colors.WHITE,
            expand=True,
            style=ft.ButtonStyle(
                text_style=ft.TextStyle(
                    size=13,
                    weight=ft.FontWeight.NORMAL
                ),
                alignment=ft.alignment.center,
                padding=ft.padding.symmetric(horizontal=8, vertical=8),
                shape=ft.RoundedRectangleBorder(radius=8)
            )
        )
        
        self.clear_btn = ft.OutlinedButton(
            "Clear Fields",
            on_click=self.clear_fields,
            expand=True,
            style=ft.ButtonStyle(
                text_style=ft.TextStyle(
                    size=13,
                    weight=ft.FontWeight.NORMAL
                ),
                alignment=ft.alignment.center,
                padding=ft.padding.symmetric(horizontal=8, vertical=8),
                shape=ft.RoundedRectangleBorder(radius=8)
            )
        )
        
        self.test_btn = ft.OutlinedButton(
            "Fill Test Data",
            on_click=self.fill_test_data,
            expand=True,
            style=ft.ButtonStyle(
                text_style=ft.TextStyle(
                    size=13,
                    weight=ft.FontWeight.NORMAL
                ),
                alignment=ft.alignment.center,
                padding=ft.padding.symmetric(horizontal=8, vertical=8),
                shape=ft.RoundedRectangleBorder(radius=8)
            )
        )
        
        self.export_btn = ft.OutlinedButton(
            "Export to Excel",
            on_click=self.export_to_excel,
            expand=True,
            style=ft.ButtonStyle(
                text_style=ft.TextStyle(
                    size=13,
                    weight=ft.FontWeight.NORMAL
                ),
                alignment=ft.alignment.center,
                padding=ft.padding.symmetric(horizontal=8, vertical=8),
                shape=ft.RoundedRectangleBorder(radius=8)
            )
        )
        
        self.remove_btn = ft.OutlinedButton(
            "Remove Selected Record",
            on_click=self.remove_record,
            expand=True,
            style=ft.ButtonStyle(
                text_style=ft.TextStyle(
                    size=13,
                    weight=ft.FontWeight.NORMAL
                ),
                alignment=ft.alignment.center,
                padding=ft.padding.symmetric(horizontal=8, vertical=8),
                shape=ft.RoundedRectangleBorder(radius=8)
            )
        )
        

        
        # QR Code preview - responsive
        self.qr_image = ft.Image(
            width=180,
            height=180,
            fit=ft.ImageFit.CONTAIN
        )
        
        # Records table - responsive with custom selection and QR display
        self.records_table = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Container(
                    content=ft.Text("Select", weight=ft.FontWeight.BOLD, size=11, text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    expand=True
                )),
                ft.DataColumn(ft.Container(
                    content=ft.Text("ID", weight=ft.FontWeight.BOLD, size=11, text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    expand=True
                )),
                ft.DataColumn(ft.Container(
                    content=ft.Text("QR", weight=ft.FontWeight.BOLD, size=11, text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    expand=True
                )),
                ft.DataColumn(ft.Container(
                    content=ft.Text("Serial Number", weight=ft.FontWeight.BOLD, size=11, text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    expand=True
                )),
                ft.DataColumn(ft.Container(
                    content=ft.Text("V-Code", weight=ft.FontWeight.BOLD, size=11, text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    expand=True
                )),
                ft.DataColumn(ft.Container(
                    content=ft.Text("DevUID", weight=ft.FontWeight.BOLD, size=11, text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    expand=True
                )),
                ft.DataColumn(ft.Container(
                    content=ft.Text("Device Name", weight=ft.FontWeight.BOLD, size=11, text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    expand=True
                )),
                ft.DataColumn(ft.Container(
                    content=ft.Text("Created (SAST)", weight=ft.FontWeight.BOLD, size=11, text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                    expand=True
                )),
            ],
            rows=[],
            horizontal_lines=ft.border.BorderSide(0.5, ft.Colors.GREY),
            column_spacing=8,
        )
        
        # Status text - properly centered in container
        self.status_text = ft.Text(
            "✅ Ready to generate QR codes",
            size=13,
            color=ft.Colors.GREEN,
            text_align=ft.TextAlign.CENTER,
            weight=ft.FontWeight.W_500
        )
        
        # Initialize selected record
        self.selected_record_id = None
    
    def build(self):
        """Build the page layout with responsive design"""
        
        # DevUID field with Get DevUID button
        devuid_row = ft.Row([
            ft.Container(content=self.devuid_field, expand=True),
            ft.Container(
                content=self.get_devuid_btn,
                width=120,
                padding=ft.padding.only(left=5)
            )
        ], spacing=5)
        
        # Input fields section - responsive with better spacing
        input_fields_section = ft.ResponsiveRow([
            ft.Container(
                content=ft.Column([
                    self.serial_field,
                    self.vcode_field,
                    devuid_row,
                    self.device_name_field,
                ], spacing=10),
                col={"sm": 12, "md": 6, "lg": 6},
                padding=8,
                alignment=ft.alignment.center
            ),
            ft.Container(
                content=ft.Column([
                    self.format_dropdown,
                    self.format_help,
                ], spacing=10),
                col={"sm": 12, "md": 6, "lg": 6},
                padding=8,
                alignment=ft.alignment.center
            ),
        ])
        
        # Main buttons section - responsive with proper column sizing for all window sizes
        buttons_section = ft.ResponsiveRow([
            ft.Container(
                content=self.generate_btn,
                col={"sm": 12, "md": 12, "lg": 3},
                padding=3,
                alignment=ft.alignment.center
            ),
            ft.Container(
                content=self.clear_btn,
                col={"sm": 6, "md": 6, "lg": 2},
                padding=3,
                alignment=ft.alignment.center
            ),
            ft.Container(
                content=self.test_btn,
                col={"sm": 6, "md": 6, "lg": 2},
                padding=3,
                alignment=ft.alignment.center
            ),
            ft.Container(
                content=self.export_btn,
                col={"sm": 6, "md": 6, "lg": 2},
                padding=3,
                alignment=ft.alignment.center
            ),
            ft.Container(
                content=self.remove_btn,
                col={"sm": 6, "md": 6, "lg": 3},
                padding=3,
                alignment=ft.alignment.center
            ),
        ])
        

        
        # Input section container
        input_section = ft.Container(
            content=ft.Column([
                input_fields_section,
                buttons_section,
                ft.Container(
                    content=self.status_text,
                    alignment=ft.alignment.center,
                    padding=ft.padding.symmetric(vertical=5),
                    expand=True
                ),
            ], spacing=15),
            padding=15,
            border=ft.border.all(1, ft.Colors.GREY_300),
            border_radius=10,
            expand=False
        )
        
        # QR Preview section - responsive
        qr_preview_section = ft.Container(
            content=ft.Column([
                ft.Container(
                    content=ft.Text("📱 QR Code Preview", size=16, weight=ft.FontWeight.BOLD, text_align=ft.TextAlign.CENTER, color=ft.Colors.BLUE_700),
                    alignment=ft.alignment.center,
                    expand=True
                ),
                ft.Container(
                    content=self.qr_image,
                    alignment=ft.alignment.center,
                    padding=10,
                    border=ft.border.all(1, ft.Colors.GREY_200),
                    border_radius=5
                ),
            ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=10),
            padding=15,
            border=ft.border.all(1, ft.Colors.GREY_300),
            border_radius=10
        )
        
        # Main input and preview row - responsive with better alignment
        main_row = ft.ResponsiveRow([
            ft.Container(
                content=input_section,
                col={"sm": 12, "md": 12, "lg": 8},
                padding=8,
                alignment=ft.alignment.center
            ),
            ft.Container(
                content=qr_preview_section,
                col={"sm": 12, "md": 12, "lg": 4},
                padding=8,
                alignment=ft.alignment.center
            ),
        ])
        
        # Records section - responsive with scrolling
        records_section = ft.Container(
            content=ft.Column([
                ft.Container(
                    content=ft.Text("📊 Recent Records", size=16, weight=ft.FontWeight.BOLD, text_align=ft.TextAlign.CENTER, color=ft.Colors.BLUE_700),
                    alignment=ft.alignment.center,
                    expand=True
                ),
                ft.Container(
                    content=ft.Row([
                        ft.Container(
                            content=self.records_table,
                            alignment=ft.alignment.center,
                            expand=True
                        )
                    ], scroll=ft.ScrollMode.AUTO, alignment=ft.MainAxisAlignment.CENTER),
                    height=400,
                    border=ft.border.all(1, ft.Colors.GREY_300),
                    border_radius=5,
                    padding=5
                )
            ], spacing=10),
            padding=15,
            border=ft.border.all(1, ft.Colors.GREY_300),
            border_radius=10,
            margin=ft.margin.only(top=10)
        )
        
        # Main layout - responsive
        main_content = ft.Column([
            self.title,
            main_row,
            records_section,
        ], spacing=15, expand=True, scroll=ft.ScrollMode.AUTO)
        
        # Add to page
        self.page.add(main_content)
    
    def on_format_change(self, e):
        """Update help text when format selection changes"""
        format_descriptions = {
            'olarm': 'Olarm (recommended) - Compatible with Olarm validation system',
            'json': 'JSON - Structured data format {"sn":"...","vc":"...","uid":"..."}',
            'csv': 'CSV - Comma-separated values: serial,verification,devuid',
            'pipe': 'Pipe - Pipe-separated values: serial|verification|devuid',
            'compact': 'Compact - Colon-separated: serial:verification:devuid',
            'labeled': 'Labeled - Human-readable with labels (old format)',
            'url': 'URL - Generic validation URL format'
        }
        selected_format = self.format_dropdown.value
        self.format_help.value = format_descriptions.get(selected_format, '')
        self.page.update()
    
    def generate_qr_code(self, e):
        """Generate QR code with input data"""
        serial_number = self.serial_field.value.strip() if self.serial_field.value else ""
        verification_code = self.vcode_field.value.strip() if self.vcode_field.value else ""
        dev_uid = self.devuid_field.value.strip() if self.devuid_field.value else ""
        device_name = self.device_name_field.value.strip() if self.device_name_field.value else ""
        format_type = self.format_dropdown.value
        
        # Validate inputs
        if not all([serial_number, verification_code, dev_uid]):
            self.show_error("Please fill in all fields")
            return
        
        try:
            # Create QR code
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_M,
                box_size=10,
                border=4,
            )
            
            # Prepare QR code data based on format type
            if format_type == "json":
                qr_data = json.dumps({
                    "sn": serial_number,
                    "vc": verification_code,
                    "uid": dev_uid
                }, separators=(',', ':'))
            elif format_type == "csv":
                qr_data = f"{serial_number},{verification_code},{dev_uid}"
            elif format_type == "pipe":
                qr_data = f"{serial_number}|{verification_code}|{dev_uid}"
            elif format_type == "labeled":
                qr_data = f"Serial Number: {serial_number}\nVerification Code: {verification_code}\nDevUID: {dev_uid}"
            elif format_type == "compact":
                qr_data = f"{serial_number}:{verification_code}:{dev_uid}"
            elif format_type == "url":
                qr_data = f"https://validate.example.com?sn={serial_number}&vc={verification_code}&uid={dev_uid}"
            elif format_type == "olarm":
                qr_data = f"https://olarm.com/o/flxr?a={serial_number},{dev_uid},{verification_code}"
            else:
                qr_data = f"https://olarm.com/o/flxr?a={serial_number},{dev_uid},{verification_code}"
            
            qr.add_data(qr_data)
            qr.make(fit=True)
            
            # Create QR code image
            qr_image = qr.make_image(fill_color="black", back_color="white")
            
            # Generate filename with timestamp (SAST)
            timestamp = datetime.now(self.sast_tz).strftime("%Y%m%d_%H%M%S")
            filename = f"qr_code_{serial_number}_{timestamp}_SAST.png"
            
            # Save QR code image
            qr_image.save(filename)
            
            # Display QR code in preview
            self.display_qr_preview(qr_image)
            
            # Save to database
            self.save_to_database(serial_number, verification_code, dev_uid, device_name, filename)
            
            # Auto-export to Excel after each new record
            try:
                self.export_to_excel_silent()
            except:
                pass
            
            # Refresh records table
            self.load_records()
            
            # Show success message
            self.show_success(f"QR code generated successfully! File: {filename}")
            
        except Exception as e:
            self.show_error(f"Failed to generate QR code: {str(e)}")
    
    def display_qr_preview(self, qr_image):
        """Display QR code in the preview area"""
        try:
            # Convert PIL image to base64 for Flet
            img_buffer = io.BytesIO()
            qr_image.save(img_buffer, format='PNG')
            img_buffer.seek(0)
            
            # Convert to base64
            img_base64 = base64.b64encode(img_buffer.read()).decode()
            
            # Update image
            self.qr_image.src_base64 = img_base64
            self.page.update()
        except Exception as e:
            print(f"Error displaying QR preview: {e}")
    
    def save_to_database(self, serial_number, verification_code, dev_uid, device_name, filename):
        """Save record to database"""
        try:
            # Find the next available ID
            self.cursor.execute('SELECT MAX(id) FROM qr_records')
            max_id = self.cursor.fetchone()[0]
            next_id = (max_id + 1) if max_id else 1
            
            # Insert with specific ID and SAST timestamp
            sast_timestamp = datetime.now(self.sast_tz).isoformat()
            self.cursor.execute('''
                INSERT INTO qr_records (id, serial_number, verification_code, dev_uid, device_name, qr_filename, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (next_id, serial_number, verification_code, dev_uid, device_name, filename, sast_timestamp))
            self.conn.commit()
            
            sast_time = datetime.now(self.sast_tz).strftime("%Y-%m-%d %H:%M:%S SAST")
            print(f"Record saved with ID: {next_id} at {sast_time}" + (f" (Device: {device_name})" if device_name else ""))
            
        except Exception as e:
            self.show_error(f"Failed to save to database: {str(e)}")
    
    def create_qr_thumbnail(self, qr_filename):
        """Create a small QR code thumbnail for table display"""
        try:
            if qr_filename and os.path.exists(qr_filename):
                # Load and resize the QR code image
                with Image.open(qr_filename) as img:
                    # Resize to medium thumbnail (doubled size)
                    img_resized = img.resize((80, 80), Image.Resampling.LANCZOS)
                    
                    # Convert to base64
                    img_buffer = io.BytesIO()
                    img_resized.save(img_buffer, format='PNG')
                    img_buffer.seek(0)
                    img_base64 = base64.b64encode(img_buffer.read()).decode()
                    
                    # Return medium image
                    return ft.Image(
                        src_base64=img_base64,
                        width=80,
                        height=80,
                        fit=ft.ImageFit.CONTAIN,
                        tooltip="Click to view full size"
                    )
            else:
                # Return placeholder text
                return ft.Text(
                    "📄",
                    size=16,
                    text_align=ft.TextAlign.CENTER,
                    tooltip="QR file not found"
                )
        except Exception as e:
            print(f"Error creating QR thumbnail: {e}")
            return ft.Text(
                "❌",
                size=16,
                text_align=ft.TextAlign.CENTER,
                tooltip="Error loading QR"
            )
    
    def load_records(self):
        """Load records from database into table"""
        try:
            self.cursor.execute('''
                SELECT id, serial_number, verification_code, dev_uid, created_at, qr_filename, device_name
                FROM qr_records
                ORDER BY created_at DESC
                LIMIT 20
            ''')
            records = self.cursor.fetchall()
            
            # Clear existing rows
            self.records_table.rows.clear()
            
            # Add records to table
            for record in records:
                try:
                    # Parse SAST timestamp and format for display
                    sast_dt = datetime.fromisoformat(record[4].replace('Z', '+02:00'))
                    created_at = sast_dt.strftime("%m-%d %H:%M SAST")
                except:
                    # Fallback for older records
                    created_at = str(record[4])[:16] + " SAST"
                    
                # Display more of the data with less truncation
                serial_display = record[1][:18] + "..." if len(record[1]) > 18 else record[1]
                devuid_display = record[3][:16] + "..." if len(record[3]) > 16 else record[3]
                device_name = record[6] if len(record) > 6 and record[6] else ""
                device_name_display = device_name[:15] + "..." if len(device_name) > 15 else device_name
                
                # Create select button for this row
                record_id = str(record[0])
                is_selected = record_id == self.selected_record_id
                
                select_btn = ft.ElevatedButton(
                    "✓" if is_selected else "○",
                    height=28,
                    width=35,
                    bgcolor=ft.Colors.GREEN if is_selected else ft.Colors.GREY_200,
                    color=ft.Colors.WHITE if is_selected else ft.Colors.BLACK,
                    on_click=lambda e, rid=record_id: self.select_record(rid),
                    style=ft.ButtonStyle(
                        text_style=ft.TextStyle(
                            size=14,
                            weight=ft.FontWeight.BOLD
                        ),
                        alignment=ft.alignment.center,
                        padding=ft.padding.all(0)
                    )
                )
                
                # Create QR code thumbnail
                qr_display = self.create_qr_thumbnail(record[5] if len(record) > 5 else None)
                
                # Create row with all columns - all text properly centered in containers
                row = ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Container(content=select_btn, alignment=ft.alignment.center, expand=True)),
                        ft.DataCell(ft.Container(
                            content=ft.Text(str(record[0]), size=10, text_align=ft.TextAlign.CENTER, weight=ft.FontWeight.BOLD),
                            alignment=ft.alignment.center,
                            expand=True
                        )),
                        ft.DataCell(ft.Container(content=qr_display, alignment=ft.alignment.center, expand=True)),
                        ft.DataCell(ft.Container(
                            content=ft.Text(serial_display, size=10, text_align=ft.TextAlign.CENTER),
                            alignment=ft.alignment.center,
                            expand=True
                        )),
                        ft.DataCell(ft.Container(
                            content=ft.Text(record[2], size=10, text_align=ft.TextAlign.CENTER),
                            alignment=ft.alignment.center,
                            expand=True
                        )),
                        ft.DataCell(ft.Container(
                            content=ft.Text(devuid_display, size=10, text_align=ft.TextAlign.CENTER),
                            alignment=ft.alignment.center,
                            expand=True
                        )),
                        ft.DataCell(ft.Container(
                            content=ft.Text(device_name_display or "-", size=10, text_align=ft.TextAlign.CENTER, italic=not device_name_display),
                            alignment=ft.alignment.center,
                            expand=True
                        )),
                        ft.DataCell(ft.Container(
                            content=ft.Text(created_at, size=10, text_align=ft.TextAlign.CENTER),
                            alignment=ft.alignment.center,
                            expand=True
                        )),
                    ],
                    color=ft.Colors.BLUE_50 if is_selected else None
                )
                
                # Store record ID as data for easy access
                row.data = record_id
                
                self.records_table.rows.append(row)
            
            self.page.update()
            
        except Exception as e:
            self.show_error(f"Failed to load records: {str(e)}")
    
    def select_record(self, record_id):
        """Select or deselect a record using custom button"""
        try:
            if self.selected_record_id == record_id:
                # Deselect if already selected
                self.selected_record_id = None
                print(f"❌ Deselected record ID: {record_id}")
                self.show_success("No record selected")
            else:
                # Select the record
                self.selected_record_id = record_id
                print(f"✅ Selected record ID: {record_id}")
                self.show_success(f"✅ Selected record ID: {record_id}. Click 'Remove Selected Record' to delete.")
            
            # Refresh the table to update button states
            self.load_records()
            
        except Exception as ex:
            print(f"❌ Error in record selection: {ex}")
            self.show_error(f"Selection error: {ex}")
    
    def get_devuid_from_device(self, e):
        """Extract DevUID from connected STM32 device using OpenOCD"""
        try:
            self.show_success("🔌 Connecting to STM32 device...")
            
            # OpenOCD command configuration
            WL_DEVUI_ADDRESS = "0x1FFF7580"
            openocd_cmd = [
                './openocd/bin/openocd',
                '-f', 'interface/stlink.cfg',
                '-f', 'target/stm32wlx.cfg',
                '-c', 'init',
                '-c', 'reset halt',
                '-c', f'mdw {WL_DEVUI_ADDRESS} 3',
                '-c', 'exit'
            ]
            
            # Run OpenOCD command with timeout
            result = subprocess.run(
                openocd_cmd,
                capture_output=True,
                text=True,
                timeout=15,  # 15 second timeout
                cwd=os.getcwd()  # Run in current directory
            )
            
            if result.returncode != 0:
                self.show_error(f"❌ OpenOCD failed: {result.stderr.strip()}")
                return
            
            # Parse the output to extract DevUID
            output_lines = result.stdout.split('\n')
            
            for line in output_lines:
                if WL_DEVUI_ADDRESS.lower() in line.lower():
                    # Split the line and extract the hex values
                    parts = line.split()
                    if len(parts) >= 4:  # Should have address and 3 hex values
                        try:
                            # Extract the hex values (skip the address)
                            hex_values = parts[1:4]
                            
                            # Combine and format the DevUID
                            # Reverse byte order for proper UID format
                            devui = hex_values[1] + hex_values[0]  # Combine first two hex values
                            
                            # Format as uppercase hex string without spaces
                            formatted_uid = devui.upper().replace('0X', '')
                            
                            # Remove any extra characters and ensure it's clean hex
                            clean_uid = ''.join(c for c in formatted_uid if c in '0123456789ABCDEF')
                            
                            if len(clean_uid) >= 16:  # Valid DevUID should be at least 16 hex chars
                                self.devuid_field.value = clean_uid
                                self.show_success(f"✅ DevUID extracted: {clean_uid}")
                                self.page.update()
                                return
                            else:
                                self.show_error(f"❌ Invalid DevUID format: {clean_uid}")
                                return
                                
                        except (IndexError, ValueError) as e:
                            self.show_error(f"❌ Error parsing DevUID: {str(e)}")
                            return
            
            self.show_error("❌ DevUID not found in OpenOCD output")
            
        except subprocess.TimeoutExpired:
            self.show_error("❌ Timeout: Device connection failed. Check ST-Link connection.")
        except FileNotFoundError:
            self.show_error("❌ OpenOCD not found. Please ensure OpenOCD is installed and available.")
        except Exception as e:
            self.show_error(f"❌ Error extracting DevUID: {str(e)}")
    
    def clear_fields(self, e):
        """Clear all input fields"""
        self.serial_field.value = ""
        self.vcode_field.value = ""
        self.devuid_field.value = ""
        self.device_name_field.value = ""
        self.qr_image.src_base64 = None
        self.status_text.value = "Fields cleared"
        self.status_text.color = ft.Colors.GREEN
        self.page.update()
    
    def fill_test_data(self, e):
        """Fill fields with test data for easy testing"""
        # Clear existing data first
        self.clear_fields(e)
        
        # Generate test data
        test_serial = f"TEST{random.randint(100000000000, 999999999999)}"
        test_vcode = f"{random.randint(100000, 999999)}"
        test_devuid = f"{''.join(random.choices('0123456789ABCDEF', k=16))}"
        
        # Fill the fields
        self.serial_field.value = test_serial
        self.vcode_field.value = test_vcode
        self.devuid_field.value = test_devuid
        
        # Set format to Olarm (default)
        self.format_dropdown.value = "olarm"
        self.on_format_change(None)
        
        self.status_text.value = "Test data filled"
        self.status_text.color = ft.Colors.BLUE
        self.page.update()
        
        sast_time = datetime.now(self.sast_tz).strftime("%Y-%m-%d %H:%M:%S SAST")
        print(f"📝 Test data filled ({sast_time}):")
        print(f"   Serial: {test_serial}")
        print(f"   V-Code: {test_vcode}")
        print(f"   DevUID: {test_devuid}")
        print(f"   Format: olarm")
    
    def remove_record(self, e):
        """Remove a selected record from the database"""
        if not self.selected_record_id:
            self.show_error("❌ No record selected! Please click the '○' button next to a record to select it first.")
            return
        
        record_id = self.selected_record_id
        
        try:
            # Get QR filename from database
            self.cursor.execute('SELECT qr_filename, serial_number FROM qr_records WHERE id = ?', (record_id,))
            result = self.cursor.fetchone()
            
            if result:
                qr_filename, serial_number = result
                
                # Delete the QR code file if it exists
                if os.path.exists(qr_filename):
                    os.remove(qr_filename)
                    print(f"Deleted QR code file: {qr_filename}")
                
                # Delete from database
                self.cursor.execute('DELETE FROM qr_records WHERE id = ?', (record_id,))
                self.conn.commit()
                
                # Clear selection
                self.selected_record_id = None
                
                # Auto-update Excel file
                try:
                    self.export_to_excel_silent()
                except:
                    pass
                
                # Refresh records table
                self.load_records()
                
                # Clear QR preview
                self.qr_image.src_base64 = None
                
                self.show_success(f"Record ID {record_id} (Serial: {serial_number}) deleted successfully!")
                
            else:
                self.show_error(f"Record ID {record_id} not found in database.")
                
        except Exception as e:
            self.show_error(f"Failed to delete record: {str(e)}")
    

    def export_to_excel(self, e):
        """Export all records to Excel with QR code images"""
        try:
            self.export_to_excel_silent()
            self.show_success("📊 Excel file updated: qr_records.xlsx")
        except Exception as e:
            self.show_error(f"Failed to export to Excel: {str(e)}")
    
    def export_to_excel_silent(self):
        """Export all records to Excel silently"""
        try:
            # Fetch all records from database
            self.cursor.execute('''
                SELECT id, serial_number, verification_code, dev_uid, device_name, qr_filename, created_at
                FROM qr_records
                ORDER BY created_at ASC
            ''')
            records = self.cursor.fetchall()
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "QR Records"
            
            # Headers
            headers = ['ID', 'Serial Number', 'Verification Code', 'DevUID', 'Device Name', 'QR Code', 'Created At (SAST)']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data rows
            for row, record in enumerate(records, 2):
                ws.cell(row=row, column=1, value=record[0])  # ID
                ws.cell(row=row, column=2, value=record[1])  # Serial Number
                ws.cell(row=row, column=3, value=record[2])  # Verification Code
                ws.cell(row=row, column=4, value=record[3])  # DevUID
                ws.cell(row=row, column=5, value=record[4] if record[4] else "")  # Device Name
                ws.cell(row=row, column=7, value=record[6])  # Created At
                
                # Add QR code image if file exists
                qr_filename = record[5]
                if os.path.exists(qr_filename):
                    try:
                        img = OpenpyxlImage(qr_filename)
                        img.width = 100
                        img.height = 100
                        ws.add_image(img, f'F{row}')
                        ws.row_dimensions[row].height = 75
                    except:
                        ws.cell(row=row, column=6, value="[QR file not found]")
                else:
                    ws.cell(row=row, column=6, value="[QR file not found]")
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 20
            
            # Save workbook
            wb.save('qr_records.xlsx')
            sast_time = datetime.now(self.sast_tz).strftime("%Y-%m-%d %H:%M:%S SAST")
            print(f"📊 Excel file updated: qr_records.xlsx at {sast_time}")
            
        except Exception as e:
            raise e
    
    def show_error(self, message):
        """Show error message"""
        self.status_text.value = message
        self.status_text.color = ft.Colors.RED
        self.page.update()
        print(f"❌ Error: {message}")
    
    def show_success(self, message):
        """Show success message"""
        self.status_text.value = message
        self.status_text.color = ft.Colors.GREEN
        self.page.update()
        print(f"✅ Success: {message}")

def main(page: ft.Page):
    """Main function to run the Flet app"""
    try:
        app = QRGeneratorApp(page)
    except Exception as e:
        print(f"Error creating app: {e}")
        page.add(ft.Text(f"Error: {e}"))

if __name__ == "__main__":
    ft.app(target=main) 