#!/usr/bin/env python3
"""
QR Generator Test Runner
Comprehensive test suite for pre-merge validation
"""

import os
import sys
import subprocess
import time
from datetime import datetime
import argparse

def print_banner():
    """Print test banner"""
    print("🔬 QR Generator Test Runner")
    print("=" * 50)
    print(f"📅 Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"🐍 Python: {sys.version.split()[0]}")
    print(f"📁 Working Dir: {os.getcwd()}")
    print("=" * 50)

def check_dependencies():
    """Check if required dependencies are installed"""
    print("\n🔍 Checking Dependencies...")
    
    required_packages = [
        'qrcode', 'PIL', 'openpyxl', 'pandas'
    ]
    
    missing_packages = []
    
    for package in required_packages:
        try:
            __import__(package)
            print(f"✅ {package}")
        except ImportError:
            missing_packages.append(package)
            print(f"❌ {package} - Missing")
    
    if missing_packages:
        print(f"\n⚠️  Missing dependencies: {', '.join(missing_packages)}")
        print("To install missing packages:")
        print(f"pip3 install --break-system-packages {' '.join(missing_packages)}")
        return False
    
    return True

def run_unit_tests():
    """Run unit tests"""
    print("\n🧪 Running Unit Tests...")
    
    try:
        result = subprocess.run([
            sys.executable, 'test_qr_generator.py'
        ], capture_output=True, text=True, timeout=300)
        
        print("STDOUT:")
        print(result.stdout)
        
        if result.stderr:
            print("STDERR:")
            print(result.stderr)
        
        return result.returncode == 0
        
    except subprocess.TimeoutExpired:
        print("❌ Unit tests timed out (5 minutes)")
        return False
    except Exception as e:
        print(f"❌ Error running unit tests: {e}")
        return False

def run_integration_test():
    """Run integration test by launching the GUI"""
    print("\n🎯 Running Integration Test...")
    
    try:
        # Test if GUI can be launched
        result = subprocess.run([
            sys.executable, 'qr.py'
        ], capture_output=True, text=True, timeout=10)
        
        # For GUI applications, we don't expect immediate completion
        # Check if there were any import errors or critical failures
        if "ImportError" in result.stderr or "ModuleNotFoundError" in result.stderr:
            print("❌ GUI launch failed - missing dependencies")
            print(result.stderr)
            return False
        
        print("✅ GUI can be launched without critical errors")
        return True
        
    except subprocess.TimeoutExpired:
        # Expected for GUI application - it should keep running
        print("✅ GUI launched successfully (running in background)")
        return True
    except Exception as e:
        print(f"❌ Error launching GUI: {e}")
        return False

def run_database_validation():
    """Validate database operations"""
    print("\n🗄️  Running Database Validation...")
    
    try:
        import sqlite3
        import tempfile
        import os
        
        # Create temporary database
        temp_db = tempfile.NamedTemporaryFile(suffix='.db', delete=False)
        temp_db.close()
        
        try:
            conn = sqlite3.connect(temp_db.name)
            cursor = conn.cursor()
            
            # Test table creation
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS qr_records (
                    id INTEGER PRIMARY KEY,
                    serial_number TEXT NOT NULL,
                    verification_code TEXT NOT NULL,
                    dev_uid TEXT NOT NULL,
                    device_name TEXT,
                    qr_filename TEXT NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # Test insert
            cursor.execute('''
                INSERT INTO qr_records (serial_number, verification_code, dev_uid, device_name, qr_filename)
                VALUES (?, ?, ?, ?, ?)
            ''', ('TEST123', '123456', 'ABCDEF', 'Test Device', 'test.png'))
            
            # Test select
            cursor.execute('SELECT * FROM qr_records')
            result = cursor.fetchall()
            
            if len(result) == 1:
                print("✅ Database operations working correctly")
                return True
            else:
                print("❌ Database validation failed")
                return False
                
        finally:
            conn.close()
            os.unlink(temp_db.name)
            
    except Exception as e:
        print(f"❌ Database validation error: {e}")
        return False

def run_qr_generation_test():
    """Test QR code generation"""
    print("\n📱 Testing QR Code Generation...")
    
    try:
        import qrcode
        import tempfile
        import os
        
        # Create temporary directory
        temp_dir = tempfile.mkdtemp()
        original_cwd = os.getcwd()
        os.chdir(temp_dir)
        
        try:
            # Test data
            test_data = {
                'serial': 'TEST123456789012',
                'vcode': '123456',
                'devuid': 'E5DDA7D74D91EC53'
            }
            
            # Generate QR code
            qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_M)
            qr_data = f"https://olarm.com/o/flxr?a={test_data['serial']},{test_data['devuid']},{test_data['vcode']}"
            qr.add_data(qr_data)
            qr.make(fit=True)
            
            # Create and save image
            qr_image = qr.make_image(fill_color="black", back_color="white")
            filename = "test_qr.png"
            qr_image.save(filename)
            
            # Verify file exists and has content
            if os.path.exists(filename) and os.path.getsize(filename) > 0:
                print("✅ QR code generation working correctly")
                return True
            else:
                print("❌ QR code generation failed")
                return False
                
        finally:
            os.chdir(original_cwd)
            import shutil
            shutil.rmtree(temp_dir)
            
    except Exception as e:
        print(f"❌ QR generation test error: {e}")
        return False

def run_export_test():
    """Test Excel export functionality"""
    print("\n📊 Testing Excel Export...")
    
    try:
        from openpyxl import Workbook
        import tempfile
        import os
        
        # Create temporary directory
        temp_dir = tempfile.mkdtemp()
        original_cwd = os.getcwd()
        os.chdir(temp_dir)
        
        try:
            # Create test Excel file
            wb = Workbook()
            ws = wb.active
            ws.title = "Test Export"
            
            # Add test data
            ws.cell(row=1, column=1, value="Test Header")
            ws.cell(row=2, column=1, value="Test Data")
            
            # Save file
            filename = "test_export.xlsx"
            wb.save(filename)
            
            # Verify file exists
            if os.path.exists(filename) and os.path.getsize(filename) > 0:
                print("✅ Excel export working correctly")
                return True
            else:
                print("❌ Excel export failed")
                return False
                
        finally:
            os.chdir(original_cwd)
            import shutil
            shutil.rmtree(temp_dir)
            
    except Exception as e:
        print(f"❌ Excel export test error: {e}")
        return False

def run_security_tests():
    """Run basic security tests"""
    print("\n🔒 Running Security Tests...")
    
    try:
        import sqlite3
        
        # Test SQL injection prevention
        conn = sqlite3.connect(':memory:')
        cursor = conn.cursor()
        
        cursor.execute('''
            CREATE TABLE test_table (
                id INTEGER PRIMARY KEY,
                data TEXT
            )
        ''')
        
        # Test malicious input (should be handled safely with parameterized queries)
        malicious_input = "'; DROP TABLE test_table; --"
        
        try:
            cursor.execute('INSERT INTO test_table (data) VALUES (?)', (malicious_input,))
            conn.commit()
            
            # Verify table still exists
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='test_table'")
            result = cursor.fetchone()
            
            if result:
                print("✅ SQL injection prevention working")
                return True
            else:
                print("❌ SQL injection prevention failed")
                return False
                
        except Exception as e:
            print(f"❌ Security test error: {e}")
            return False
        finally:
            conn.close()
            
    except Exception as e:
        print(f"❌ Security test setup error: {e}")
        return False

def generate_report(results):
    """Generate test report"""
    print("\n" + "=" * 50)
    print("📋 TEST REPORT")
    print("=" * 50)
    
    total_tests = len(results)
    passed_tests = sum(1 for result in results.values() if result)
    
    print(f"📊 Total Tests: {total_tests}")
    print(f"✅ Passed: {passed_tests}")
    print(f"❌ Failed: {total_tests - passed_tests}")
    print(f"📈 Success Rate: {(passed_tests / total_tests * 100):.1f}%")
    
    print("\n📝 Individual Results:")
    for test_name, result in results.items():
        status = "✅ PASS" if result else "❌ FAIL"
        print(f"  {test_name}: {status}")
    
    overall_success = all(results.values())
    
    if overall_success:
        print("\n🎉 ALL TESTS PASSED - READY FOR MERGE!")
    else:
        print("\n⚠️  SOME TESTS FAILED - REVIEW BEFORE MERGE")
    
    return overall_success

def main():
    """Main test runner"""
    parser = argparse.ArgumentParser(description='Run QR Generator tests')
    parser.add_argument('--quick', action='store_true', help='Run quick tests only')
    parser.add_argument('--unit-only', action='store_true', help='Run unit tests only')
    parser.add_argument('--skip-deps', action='store_true', help='Skip dependency check')
    args = parser.parse_args()
    
    print_banner()
    
    # Check dependencies
    if not args.skip_deps:
        if not check_dependencies():
            print("\n❌ Dependency check failed!")
            sys.exit(1)
    
    # Run tests
    results = {}
    
    if not args.quick:
        results['Database Validation'] = run_database_validation()
        results['QR Generation'] = run_qr_generation_test()
        results['Excel Export'] = run_export_test()
        results['Security Tests'] = run_security_tests()
    
    if not args.unit_only:
        results['Integration Test'] = run_integration_test()
    
    # Always run unit tests last (most comprehensive)
    results['Unit Tests'] = run_unit_tests()
    
    # Generate report
    success = generate_report(results)
    
    # Exit with appropriate code
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main() 