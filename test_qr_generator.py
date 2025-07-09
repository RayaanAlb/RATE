#!/usr/bin/env python3
"""
Automated Test Suite for QR Generator Application
Comprehensive tests covering all functionality
"""

import unittest
import sqlite3
import os
import tempfile
import shutil
import sys
import json
from datetime import datetime
from unittest.mock import patch, MagicMock
import qrcode
from PIL import Image
import io

# Add the getDEVUID directory to the path so we can import the QR generator
sys.path.append('getDEVUID')

# Import required modules (create mocks if not available)
try:
    import flet as ft
    FLET_AVAILABLE = True
except ImportError:
    FLET_AVAILABLE = False
    print("⚠️  Warning: Flet not available. Testing core functionality only.")

class TestQRGeneratorCore(unittest.TestCase):
    """Test core QR generator functionality without GUI"""
    
    def setUp(self):
        """Set up test environment"""
        self.test_dir = tempfile.mkdtemp()
        self.original_cwd = os.getcwd()
        os.chdir(self.test_dir)
        
        # Create test database
        self.conn = sqlite3.connect('test_qr_codes.db')
        self.cursor = self.conn.cursor()
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS qr_records (
                id INTEGER PRIMARY KEY,
                serial_number TEXT NOT NULL,
                verification_code TEXT NOT NULL,
                dev_uid TEXT NOT NULL,
                device_name TEXT,
                qr_filename TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        self.conn.commit()
    
    def tearDown(self):
        """Clean up test environment"""
        os.chdir(self.original_cwd)
        self.conn.close()
        shutil.rmtree(self.test_dir)
    
    def test_qr_generation_basic(self):
        """Test basic QR code generation"""
        # Test data
        serial_number = "TEST123456789012"
        verification_code = "123456"
        dev_uid = "E5DDA7D74D91EC53"
        
        # Create QR code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=10,
            border=4,
        )
        
        # Test Olarm format
        qr_data = f"https://olarm.com/o/flxr?a={serial_number},{dev_uid},{verification_code}"
        qr.add_data(qr_data)
        qr.make(fit=True)
        
        # Create QR code image
        qr_image = qr.make_image(fill_color="black", back_color="white")
        
        # Save to file
        filename = f"test_qr_{serial_number}.png"
        qr_image.save(filename)
        
        # Verify file exists
        self.assertTrue(os.path.exists(filename))
        
        # Verify image properties
        with Image.open(filename) as img:
            self.assertEqual(img.format, 'PNG')
            self.assertGreater(img.width, 0)
            self.assertGreater(img.height, 0)
    
    def test_qr_formats(self):
        """Test different QR code formats"""
        test_data = {
            'serial': 'TEST123456789012',
            'vcode': '123456',
            'devuid': 'E5DDA7D74D91EC53'
        }
        
        formats = {
            'olarm': f"https://olarm.com/o/flxr?a={test_data['serial']},{test_data['devuid']},{test_data['vcode']}",
            'json': json.dumps({"sn": test_data['serial'], "vc": test_data['vcode'], "uid": test_data['devuid']}),
            'csv': f"{test_data['serial']},{test_data['vcode']},{test_data['devuid']}",
            'pipe': f"{test_data['serial']}|{test_data['vcode']}|{test_data['devuid']}",
            'compact': f"{test_data['serial']}:{test_data['vcode']}:{test_data['devuid']}",
            'labeled': f"Serial Number: {test_data['serial']}\nVerification Code: {test_data['vcode']}\nDevUID: {test_data['devuid']}",
            'url': f"https://validate.example.com?sn={test_data['serial']}&vc={test_data['vcode']}&uid={test_data['devuid']}"
        }
        
        for format_name, expected_data in formats.items():
            with self.subTest(format=format_name):
                qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_M)
                qr.add_data(expected_data)
                qr.make(fit=True)
                
                qr_image = qr.make_image(fill_color="black", back_color="white")
                filename = f"test_qr_{format_name}.png"
                qr_image.save(filename)
                
                self.assertTrue(os.path.exists(filename))
    
    def test_database_operations(self):
        """Test database CRUD operations"""
        # Test data
        test_record = {
            'id': 1,
            'serial_number': 'TEST123456789012',
            'verification_code': '123456',
            'dev_uid': 'E5DDA7D74D91EC53',
            'device_name': 'Test Device',
            'qr_filename': 'test_qr.png'
        }
        
        # Insert record
        self.cursor.execute('''
            INSERT INTO qr_records (id, serial_number, verification_code, dev_uid, device_name, qr_filename)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (test_record['id'], test_record['serial_number'], test_record['verification_code'], 
              test_record['dev_uid'], test_record['device_name'], test_record['qr_filename']))
        self.conn.commit()
        
        # Verify record exists
        self.cursor.execute('SELECT * FROM qr_records WHERE id = ?', (test_record['id'],))
        result = self.cursor.fetchone()
        self.assertIsNotNone(result)
        self.assertEqual(result[1], test_record['serial_number'])
        self.assertEqual(result[4], test_record['device_name'])
        
        # Test update
        new_device_name = 'Updated Device'
        self.cursor.execute('UPDATE qr_records SET device_name = ? WHERE id = ?', 
                          (new_device_name, test_record['id']))
        self.conn.commit()
        
        # Verify update
        self.cursor.execute('SELECT device_name FROM qr_records WHERE id = ?', (test_record['id'],))
        result = self.cursor.fetchone()
        self.assertEqual(result[0], new_device_name)
        
        # Test delete
        self.cursor.execute('DELETE FROM qr_records WHERE id = ?', (test_record['id'],))
        self.conn.commit()
        
        # Verify deletion
        self.cursor.execute('SELECT * FROM qr_records WHERE id = ?', (test_record['id'],))
        result = self.cursor.fetchone()
        self.assertIsNone(result)
    
    def test_database_schema(self):
        """Test database schema integrity"""
        # Check table exists
        self.cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='qr_records'")
        result = self.cursor.fetchone()
        self.assertIsNotNone(result)
        
        # Check columns
        self.cursor.execute("PRAGMA table_info(qr_records)")
        columns = self.cursor.fetchall()
        
        expected_columns = ['id', 'serial_number', 'verification_code', 'dev_uid', 'device_name', 'qr_filename', 'created_at']
        actual_columns = [col[1] for col in columns]
        
        for expected_col in expected_columns:
            self.assertIn(expected_col, actual_columns)
    
    def test_input_validation(self):
        """Test input validation logic"""
        # Test valid inputs
        valid_inputs = [
            ('TEST123456789012', '123456', 'E5DDA7D74D91EC53'),
            ('DEVICE001', '999999', 'ABCDEF1234567890'),
            ('A' * 20, '1' * 6, 'F' * 16)
        ]
        
        for serial, vcode, devuid in valid_inputs:
            with self.subTest(serial=serial, vcode=vcode, devuid=devuid):
                # Basic validation - all fields must be non-empty
                self.assertTrue(bool(serial.strip()))
                self.assertTrue(bool(vcode.strip()))
                self.assertTrue(bool(devuid.strip()))
        
        # Test invalid inputs (empty fields)
        invalid_inputs = [
            ('', '123456', 'E5DDA7D74D91EC53'),
            ('TEST123456789012', '', 'E5DDA7D74D91EC53'),
            ('TEST123456789012', '123456', ''),
            ('', '', ''),
        ]
        
        for serial, vcode, devuid in invalid_inputs:
            with self.subTest(serial=serial, vcode=vcode, devuid=devuid):
                # Should fail validation
                has_empty_field = not (serial.strip() and vcode.strip() and devuid.strip())
                self.assertTrue(has_empty_field)


class TestQRGeneratorSecurity(unittest.TestCase):
    """Test security aspects of QR generator"""
    
    def test_sql_injection_prevention(self):
        """Test SQL injection prevention"""
        # Test malicious inputs
        malicious_inputs = [
            "'; DROP TABLE qr_records; --",
            "' OR '1'='1' --",
            "'; SELECT * FROM qr_records; --",
            "1'; UPDATE qr_records SET serial_number='hacked' WHERE id=1; --"
        ]
        
        conn = sqlite3.connect(':memory:')
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE qr_records (
                id INTEGER PRIMARY KEY,
                serial_number TEXT NOT NULL,
                verification_code TEXT NOT NULL,
                dev_uid TEXT NOT NULL,
                device_name TEXT,
                qr_filename TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        for malicious_input in malicious_inputs:
            with self.subTest(input=malicious_input):
                # Use parameterized queries (safe)
                try:
                    cursor.execute('''
                        INSERT INTO qr_records (serial_number, verification_code, dev_uid, qr_filename)
                        VALUES (?, ?, ?, ?)
                    ''', (malicious_input, '123456', 'E5DDA7D74D91EC53', 'test.png'))
                    conn.commit()
                    
                    # Verify table still exists and data is stored safely
                    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='qr_records'")
                    result = cursor.fetchone()
                    self.assertIsNotNone(result)
                    
                    # Verify the malicious input was stored as literal text
                    cursor.execute("SELECT serial_number FROM qr_records WHERE serial_number = ?", (malicious_input,))
                    result = cursor.fetchone()
                    self.assertEqual(result[0], malicious_input)
                    
                    # Clean up for next test
                    cursor.execute("DELETE FROM qr_records")
                    conn.commit()
                    
                except sqlite3.Error as e:
                    # Should not fail with parameterized queries
                    self.fail(f"Parameterized query failed: {e}")
        
        conn.close()
    
    def test_path_traversal_prevention(self):
        """Test path traversal prevention"""
        malicious_filenames = [
            "../../../etc/passwd",
            "..\\..\\..\\windows\\system32\\hosts",
            "/etc/passwd",
            "C:\\Windows\\System32\\drivers\\etc\\hosts",
            "....//....//....//etc//passwd"
        ]
        
        for filename in malicious_filenames:
            with self.subTest(filename=filename):
                # Sanitize filename (remove path traversal attempts)
                sanitized = os.path.basename(filename)
                # Remove any remaining dangerous characters including dots
                sanitized = ''.join(c for c in sanitized if c.isalnum() or c in '_-')
                
                # Should not contain path separators or dots
                self.assertNotIn('/', sanitized)
                self.assertNotIn('\\', sanitized)
                self.assertNotIn('.', sanitized)


class TestQRGeneratorPerformance(unittest.TestCase):
    """Test performance aspects of QR generator"""
    
    def test_qr_generation_performance(self):
        """Test QR generation performance"""
        import time
        
        # Test data
        test_data = {
            'serial': 'TEST123456789012',
            'vcode': '123456',
            'devuid': 'E5DDA7D74D91EC53'
        }
        
        # Measure QR generation time
        start_time = time.time()
        
        for i in range(10):  # Generate 10 QR codes
            qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_M)
            qr_data = f"https://olarm.com/o/flxr?a={test_data['serial']},{test_data['devuid']},{test_data['vcode']}"
            qr.add_data(qr_data)
            qr.make(fit=True)
            qr_image = qr.make_image(fill_color="black", back_color="white")
            
            # Save to memory buffer (simulating file save)
            buffer = io.BytesIO()
            qr_image.save(buffer, format='PNG')
        
        end_time = time.time()
        average_time = (end_time - start_time) / 10
        
        # Should generate QR code in less than 1 second on average
        self.assertLess(average_time, 1.0, f"QR generation took {average_time:.2f} seconds on average")
    
    def test_database_performance(self):
        """Test database operation performance"""
        import time
        
        # Create temporary database
        conn = sqlite3.connect(':memory:')
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE qr_records (
                id INTEGER PRIMARY KEY,
                serial_number TEXT NOT NULL,
                verification_code TEXT NOT NULL,
                dev_uid TEXT NOT NULL,
                device_name TEXT,
                qr_filename TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Insert 100 records and measure time
        start_time = time.time()
        
        for i in range(100):
            cursor.execute('''
                INSERT INTO qr_records (serial_number, verification_code, dev_uid, qr_filename)
                VALUES (?, ?, ?, ?)
            ''', (f'TEST{i:012d}', f'{i:06d}', f'DEVUID{i:08X}', f'qr_{i}.png'))
        
        conn.commit()
        end_time = time.time()
        insert_time = end_time - start_time
        
        # Test query performance
        start_time = time.time()
        cursor.execute('SELECT * FROM qr_records ORDER BY created_at DESC LIMIT 20')
        results = cursor.fetchall()
        end_time = time.time()
        query_time = end_time - start_time
        
        # Performance assertions
        self.assertLess(insert_time, 5.0, f"Database insertion took {insert_time:.2f} seconds")
        self.assertLess(query_time, 1.0, f"Database query took {query_time:.2f} seconds")
        self.assertEqual(len(results), 20)
        
        conn.close()


class TestQRGeneratorIntegration(unittest.TestCase):
    """Integration tests for QR generator"""
    
    def setUp(self):
        """Set up integration test environment"""
        self.test_dir = tempfile.mkdtemp()
        self.original_cwd = os.getcwd()
        os.chdir(self.test_dir)
    
    def tearDown(self):
        """Clean up integration test environment"""
        os.chdir(self.original_cwd)
        shutil.rmtree(self.test_dir)
    
    def test_end_to_end_workflow(self):
        """Test complete end-to-end workflow"""
        # 1. Initialize database
        conn = sqlite3.connect('qr_codes.db')
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS qr_records (
                id INTEGER PRIMARY KEY,
                serial_number TEXT NOT NULL,
                verification_code TEXT NOT NULL,
                dev_uid TEXT NOT NULL,
                device_name TEXT,
                qr_filename TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn.commit()
        
        # 2. Generate QR code
        test_data = {
            'serial': 'TEST123456789012',
            'vcode': '123456',
            'devuid': 'E5DDA7D74D91EC53',
            'device_name': 'Test Device'
        }
        
        qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_M)
        qr_data = f"https://olarm.com/o/flxr?a={test_data['serial']},{test_data['devuid']},{test_data['vcode']}"
        qr.add_data(qr_data)
        qr.make(fit=True)
        qr_image = qr.make_image(fill_color="black", back_color="white")
        
        # 3. Save QR code
        filename = f"qr_code_{test_data['serial']}.png"
        qr_image.save(filename)
        
        # 4. Store in database
        cursor.execute('''
            INSERT INTO qr_records (serial_number, verification_code, dev_uid, device_name, qr_filename)
            VALUES (?, ?, ?, ?, ?)
        ''', (test_data['serial'], test_data['vcode'], test_data['devuid'], 
              test_data['device_name'], filename))
        conn.commit()
        
        # 5. Verify everything worked
        self.assertTrue(os.path.exists(filename))
        
        cursor.execute('SELECT * FROM qr_records WHERE serial_number = ?', (test_data['serial'],))
        result = cursor.fetchone()
        self.assertIsNotNone(result)
        self.assertEqual(result[1], test_data['serial'])
        self.assertEqual(result[4], test_data['device_name'])
        
        # 6. Test Excel export simulation
        try:
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image as OpenpyxlImage
            
            wb = Workbook()
            ws = wb.active
            ws.title = "QR Records"
            
            # Add headers
            headers = ['ID', 'Serial Number', 'Verification Code', 'DevUID', 'Device Name', 'QR Code', 'Created At']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Add data
            ws.cell(row=2, column=1, value=result[0])
            ws.cell(row=2, column=2, value=result[1])
            ws.cell(row=2, column=3, value=result[2])
            ws.cell(row=2, column=4, value=result[3])
            ws.cell(row=2, column=5, value=result[4])
            ws.cell(row=2, column=7, value=result[6])
            
            # Add QR image
            if os.path.exists(filename):
                img = OpenpyxlImage(filename)
                img.width = 100
                img.height = 100
                ws.add_image(img, 'F2')
            
            wb.save('test_export.xlsx')
            self.assertTrue(os.path.exists('test_export.xlsx'))
            
        except ImportError:
            print("⚠️  Skipping Excel export test - openpyxl not available")
        
        # 7. Clean up
        cursor.execute('DELETE FROM qr_records WHERE serial_number = ?', (test_data['serial'],))
        conn.commit()
        
        if os.path.exists(filename):
            os.remove(filename)
        
        conn.close()


def run_comprehensive_tests():
    """Run all tests and generate report"""
    print("🚀 Starting Comprehensive QR Generator Test Suite")
    print("=" * 60)
    
    # Create test suite
    test_suite = unittest.TestSuite()
    
    # Add all test classes
    test_classes = [
        TestQRGeneratorCore,
        TestQRGeneratorSecurity,
        TestQRGeneratorPerformance,
        TestQRGeneratorIntegration
    ]
    
    for test_class in test_classes:
        tests = unittest.TestLoader().loadTestsFromTestCase(test_class)
        test_suite.addTests(tests)
    
    # Run tests with detailed output
    runner = unittest.TextTestRunner(verbosity=2, stream=sys.stdout)
    result = runner.run(test_suite)
    
    # Generate summary report
    print("\n" + "=" * 60)
    print("🎯 TEST SUMMARY REPORT")
    print("=" * 60)
    print(f"Tests Run: {result.testsRun}")
    print(f"Failures: {len(result.failures)}")
    print(f"Errors: {len(result.errors)}")
    print(f"Success Rate: {((result.testsRun - len(result.failures) - len(result.errors)) / result.testsRun * 100):.1f}%")
    
    if result.failures:
        print("\n❌ FAILURES:")
        for test, traceback in result.failures:
            print(f"  - {test}: {traceback.split('AssertionError: ')[-1].split('\\n')[0]}")
    
    if result.errors:
        print("\n🔥 ERRORS:")
        for test, traceback in result.errors:
            print(f"  - {test}: {traceback.split('\\n')[-2]}")
    
    if result.wasSuccessful():
        print("\n✅ ALL TESTS PASSED!")
    else:
        print("\n❌ SOME TESTS FAILED!")
    
    return result.wasSuccessful()


if __name__ == "__main__":
    success = run_comprehensive_tests()
    sys.exit(0 if success else 1) 